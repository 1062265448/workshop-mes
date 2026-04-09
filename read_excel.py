import openpyxl
import os

# 使用 UTF-8 编码读取文件路径
excel_path = r"D:\2026\报表管理和服务器优化改造\成品车间报表\初始表\2026 年 1 月动态表 (2).xlsx"

print(f"文件路径：{excel_path}")
print(f"文件存在：{os.path.exists(excel_path)}")

if os.path.exists(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    print(f"\n工作表数量：{len(wb.sheetnames)}")
    print(f"工作表名称：{wb.sheetnames}")
    
    ws = wb.active
    print(f"\n当前工作表：{ws.title}")
    print(f"总行数：{ws.max_row}")
    print(f"总列数：{ws.max_column}")
    
    print("\n=== 表头 (第 1 行) ===")
    headers = []
    for col in range(1, min(ws.max_column + 1, 30)):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            headers.append(cell_value)
            print(f"列{col}: {cell_value}")
    
    print("\n=== 前 5 行数据 ===")
    for row in range(2, min(7, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(len(headers) + 1, 30)):
            row_data.append(ws.cell(row=row, column=col).value)
        print(f"行{row}: {row_data}")
    
    wb.close()
else:
    print("文件不存在！")
    print("\n查找类似文件:")
    import glob
    files = glob.glob(r"D:\2026\报表管理和服务器优化改造\成品车间报表\初始表\*.xlsx")
    for f in files[:10]:
        print(f"  {f}")
